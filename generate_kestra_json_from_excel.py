#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère un JSON Kestra déclaratif complet depuis le fichier Excel GEA.

Usage :
  python generate_kestra_json_from_excel.py modele_kestra_provisioning_gea_json_ready_v4.xlsx caa-data-dev kestra-config-caa_data_dev.json
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Module Python manquant: openpyxl. Installation: pip install openpyxl") from exc


def is_active(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"true", "vrai", "yes", "oui", "1", "x", "y"}


def split_list(value):
    if value is None:
        return []
    return [item.strip() for item in str(value).replace(";", ",").split(",") if item.strip()]


def rows(ws):
    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    result = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = excel_row[idx] if idx < len(excel_row) else None
            row[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            result.append(row)
    return result


def read_sheet(workbook, name):
    if name not in workbook.sheetnames:
        raise SystemExit(f"Onglet manquant dans l'Excel: {name}")
    return rows(workbook[name])


def build_config(excel_path: str, tenant_id: str):
    wb = load_workbook(excel_path, data_only=True)

    tenants = [
        r for r in read_sheet(wb, "01_Tenants")
        if is_active(r.get("active")) and r.get("tenant_id") == tenant_id
    ]
    if not tenants:
        raise SystemExit(f"Tenant introuvable ou inactif dans 01_Tenants: {tenant_id}")

    tenant_row = tenants[0]
    tenant = {
        "tenant_id": tenant_row.get("tenant_id"),
        "label": tenant_row.get("label") or tenant_row.get("tenant_id")
    }
    if tenant_row.get("existing") not in (None, ""):
        tenant["existing"] = is_active(tenant_row.get("existing"))

    namespaces = []
    for r in read_sheet(wb, "03_Namespaces"):
        if is_active(r.get("active")) and r.get("tenant_id") == tenant_id:
            namespaces.append({
                "namespace_id": r.get("namespace_id"),
                "description": r.get("description") or ""
            })

    role_rows = {
        r.get("role_name"): r
        for r in read_sheet(wb, "05_Roles_Kestra")
        if is_active(r.get("active")) and r.get("role_name")
    }

    permissions_by_role = {}
    for r in read_sheet(wb, "06_Role_Permissions"):
        if is_active(r.get("active")) and r.get("role_name") and r.get("type"):
            permissions_by_role.setdefault(r.get("role_name"), []).append({
                "type": r.get("type"),
                "permissions": split_list(r.get("permissions"))
            })

    roles = []
    for role_name, role_row in role_rows.items():
        roles.append({
            "name": role_name,
            "description": role_row.get("description") or "",
            "permissions": permissions_by_role.get(role_name, [])
        })

    bindings = []
    referenced_groups = set()

    for r in read_sheet(wb, "08_Bindings"):
        if not is_active(r.get("active")) or r.get("tenant_id") != tenant_id:
            continue

        roles_for_binding = split_list(r.get("roles"))
        if not roles_for_binding:
            continue

        binding = {"roles": roles_for_binding}
        subject_type = str(r.get("subject_type") or "").strip().upper()

        if subject_type == "GROUP":
            subject_name = r.get("subject_name")
            binding["group"] = subject_name
            referenced_groups.add(subject_name)
        elif subject_type == "SERVICE_ACCOUNT":
            binding["service_account"] = r.get("subject_name")
        else:
            continue

        namespace = r.get("namespace") or r.get("json_namespace")
        if namespace not in (None, ""):
            binding["namespace"] = namespace

        bindings.append(binding)

    group_rows = {
        r.get("group_name"): r
        for r in read_sheet(wb, "04_Groupes_AD")
        if is_active(r.get("active")) and r.get("group_name")
    }

    groups = []
    for group_name in sorted(referenced_groups):
        group_row = group_rows.get(group_name, {})
        group = {
            "name": group_name,
            "description": group_row.get("description") or ""
        }
        if group_row.get("namespace") not in (None, ""):
            group["namespace"] = group_row.get("namespace")
        groups.append(group)

    service_accounts = []
    for r in read_sheet(wb, "07_Service_Accounts"):
        if is_active(r.get("active")) and r.get("tenant_id") in (tenant_id, None, ""):
            service_accounts.append({
                "name": r.get("name"),
                "description": r.get("description") or "",
                "generate_token": is_active(r.get("generate_token")),
                "token_description": r.get("token_description") or "",
                "token_max_age": r.get("token_max_age") or "P365D"
            })

    return {
        "tenant": tenant,
        "namespaces": namespaces,
        "roles": roles,
        "groups": groups,
        "service_accounts": service_accounts,
        "bindings": bindings
    }


def main():
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: python generate_kestra_json_from_excel.py "
            "<fichier_excel.xlsx> <tenant_id> <fichier_sortie.json>"
        )

    excel_path = sys.argv[1]
    tenant_id = sys.argv[2]
    output_path = sys.argv[3]

    if not Path(excel_path).exists():
        raise SystemExit(f"Fichier Excel introuvable: {excel_path}")

    config = build_config(excel_path, tenant_id)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"JSON généré: {output_path}")
    print(f"Tenant: {tenant_id}")
    print(f"Namespaces: {len(config['namespaces'])}")
    print(f"Roles: {len(config['roles'])}")
    print(f"Groups: {len(config['groups'])}")
    print(f"Service accounts: {len(config['service_accounts'])}")
    print(f"Bindings: {len(config['bindings'])}")


if __name__ == "__main__":
    main()
