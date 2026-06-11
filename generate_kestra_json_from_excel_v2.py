#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère un JSON Kestra déclaratif complet depuis le fichier Excel GEA.

Usage :
  python generate_kestra_json_from_excel_v2.py modele_kestra_provisioning_gea_json_ready_v4.xlsx caa-data-dev output-dev.json

Points corrigés :
- accepte caa-data-dev ou caa_data_dev si l'Excel utilise l'autre convention ;
- affiche les tenants disponibles si le tenant demandé n'est pas trouvé ;
- considère une ligne renseignée comme active par défaut si la colonne active est vide
  (utile si Excel affiche une case cochée mais que le booléen n'est pas lisible par openpyxl) ;
- permet de désactiver explicitement une ligne avec FALSE / faux / non / 0 / inactive.
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Module Python manquant: openpyxl. Installation: pip install openpyxl") from exc

FALSE_VALUES = {"false", "faux", "no", "non", "0", "n", "inactive", "inactif", "disabled", "désactivé", "desactive"}
TRUE_VALUES = {"true", "vrai", "yes", "oui", "1", "x", "y", "active", "actif", "enabled"}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def canon(value):
    """Normalisation tolérante pour comparer caa-data-dev et caa_data_dev."""
    return clean(value).lower().replace("_", "-")


def is_active(value, default=True) -> bool:
    """Active par défaut si la cellule est vide, car les cases Excel peuvent être non lisibles."""
    if isinstance(value, bool):
        return value
    if value is None or clean(value) == "":
        return default
    if isinstance(value, (int, float)):
        return value != 0
    normalized = clean(value).lower()
    if normalized in FALSE_VALUES:
        return False
    if normalized in TRUE_VALUES:
        return True
    # Si valeur non standard, on conserve la ligne par défaut.
    return default


def split_list(value):
    if value is None:
        return []
    return [item.strip() for item in str(value).replace(";", ",").split(",") if item.strip()]


def rows(ws):
    raw_headers = [cell.value for cell in ws[1]]
    headers = [clean(h) for h in raw_headers]
    result = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = excel_row[idx] if idx < len(excel_row) else None
            row[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            result.append(row)
    return result


def read_sheet(workbook, name):
    if name not in workbook.sheetnames:
        raise SystemExit(f"Onglet manquant dans l'Excel: {name}")
    return rows(workbook[name])


def tenant_matches(row_tenant_id, requested_tenant_id):
    # Match exact ou tolérant entre _ et -.
    return clean(row_tenant_id) == clean(requested_tenant_id) or canon(row_tenant_id) == canon(requested_tenant_id)


def row_tenant_in_scope(row_tenant_id, requested_tenant_id):
    """Certaines lignes peuvent être globales si tenant_id est vide, ALL ou *."""
    value = clean(row_tenant_id)
    if value == "" or value.upper() in {"ALL", "*", "GLOBAL"}:
        return True
    return tenant_matches(value, requested_tenant_id)


def build_config(excel_path: str, requested_tenant_id: str):
    wb = load_workbook(excel_path, data_only=True)

    tenant_rows = read_sheet(wb, "01_Tenants")
    tenants = [
        r for r in tenant_rows
        if is_active(r.get("active"), default=True) and tenant_matches(r.get("tenant_id"), requested_tenant_id)
    ]

    if not tenants:
        available = []
        for r in tenant_rows:
            tid = clean(r.get("tenant_id"))
            if not tid:
                continue
            active_raw = r.get("active")
            active_status = is_active(active_raw, default=True)
            available.append(f"- {tid} | active={active_status} | valeur_active_cellule={active_raw!r}")
        msg = [f"Tenant introuvable ou inactif dans 01_Tenants: {requested_tenant_id}"]
        if available:
            msg.append("Tenants visibles dans l'Excel :")
            msg.extend(available)
        raise SystemExit("\n".join(msg))

    tenant_row = tenants[0]
    effective_tenant_id = clean(tenant_row.get("tenant_id"))
    tenant = {
        "tenant_id": effective_tenant_id,
        "label": clean(tenant_row.get("label")) or effective_tenant_id
    }
    if tenant_row.get("existing") not in (None, ""):
        tenant["existing"] = is_active(tenant_row.get("existing"), default=False)

    namespaces = []
    for r in read_sheet(wb, "03_Namespaces"):
        if is_active(r.get("active"), default=True) and row_tenant_in_scope(r.get("tenant_id"), effective_tenant_id):
            namespace_id = clean(r.get("namespace_id"))
            if namespace_id:
                namespaces.append({
                    "namespace_id": namespace_id,
                    "description": clean(r.get("description"))
                })

    role_rows = {
        clean(r.get("role_name")): r
        for r in read_sheet(wb, "05_Roles_Kestra")
        if is_active(r.get("active"), default=True) and clean(r.get("role_name"))
    }

    permissions_by_role = {}
    for r in read_sheet(wb, "06_Role_Permissions"):
        if is_active(r.get("active"), default=True) and clean(r.get("role_name")) and clean(r.get("type")):
            permissions_by_role.setdefault(clean(r.get("role_name")), []).append({
                "type": clean(r.get("type")),
                "permissions": split_list(r.get("permissions"))
            })

    roles = []
    for role_name, role_row in role_rows.items():
        roles.append({
            "name": role_name,
            "description": clean(role_row.get("description")),
            "permissions": permissions_by_role.get(role_name, [])
        })

    bindings = []
    referenced_groups = set()

    for r in read_sheet(wb, "08_Bindings"):
        if not is_active(r.get("active"), default=True) or not row_tenant_in_scope(r.get("tenant_id"), effective_tenant_id):
            continue

        roles_for_binding = split_list(r.get("roles"))
        if not roles_for_binding:
            continue

        binding = {"roles": roles_for_binding}
        subject_type = clean(r.get("subject_type")).upper()
        subject_name = clean(r.get("subject_name"))

        if subject_type == "GROUP":
            if not subject_name:
                continue
            binding["group"] = subject_name
            referenced_groups.add(subject_name)
        elif subject_type == "SERVICE_ACCOUNT":
            if not subject_name:
                continue
            binding["service_account"] = subject_name
        else:
            continue

        namespace = clean(r.get("namespace")) or clean(r.get("json_namespace"))
        if namespace:
            binding["namespace"] = namespace

        bindings.append(binding)

    group_rows = {
        clean(r.get("group_name")): r
        for r in read_sheet(wb, "04_Groupes_AD")
        if is_active(r.get("active"), default=True) and clean(r.get("group_name"))
    }

    groups = []
    for group_name in sorted(referenced_groups):
        group_row = group_rows.get(group_name, {})
        group = {
            "name": group_name,
            "description": clean(group_row.get("description"))
        }
        namespace = clean(group_row.get("namespace"))
        if namespace:
            group["namespace"] = namespace
        groups.append(group)

    service_accounts = []
    for r in read_sheet(wb, "07_Service_Accounts"):
        if is_active(r.get("active"), default=True) and row_tenant_in_scope(r.get("tenant_id"), effective_tenant_id):
            name = clean(r.get("name"))
            if not name:
                continue
            service_accounts.append({
                "name": name,
                "description": clean(r.get("description")),
                "generate_token": is_active(r.get("generate_token"), default=False),
                "token_description": clean(r.get("token_description")),
                "token_max_age": clean(r.get("token_max_age")) or "P365D"
            })

    return {
        "tenant": tenant,
        "namespaces": namespaces,
        "roles": roles,
        "groups": groups,
        "service_accounts": service_accounts,
        "bindings": bindings
    }


def main():
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: python generate_kestra_json_from_excel_v2.py "
            "<fichier_excel.xlsx> <tenant_id> <fichier_sortie.json>"
        )

    excel_path = sys.argv[1]
    tenant_id = sys.argv[2]
    output_path = sys.argv[3]

    if not Path(excel_path).exists():
        raise SystemExit(f"Fichier Excel introuvable: {excel_path}")

    config = build_config(excel_path, tenant_id)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"JSON généré: {output_path}")
    print(f"Tenant demandé: {tenant_id}")
    print(f"Tenant JSON: {config['tenant']['tenant_id']}")
    print(f"Namespaces: {len(config['namespaces'])}")
    print(f"Roles: {len(config['roles'])}")
    print(f"Groups: {len(config['groups'])}")
    print(f"Service accounts: {len(config['service_accounts'])}")
    print(f"Bindings: {len(config['bindings'])}")


if __name__ == "__main__":
    main()
